import pandas as pd
from pymongo import MongoClient

# Step 1: Load CSV into a DataFrame
df1 = pd.read_csv("commuter_data.csv") 
df2 = pd.read_csv("routes_data.csv") 

# Step 2: Connect to MongoDB
client = MongoClient("mongodb://localhost:27017/")
db = client["dwh_project"]         
collection1 = db["commuters"]   
collection2 = db["routes"]   

# Step 3: Convert DataFrame to list of dictionaries
data1 = df1.to_dict(orient="records")
data2 = df2.to_dict(orient="records")
# Step 4: Insert into MongoDB
collection1.insert_many(data1)
collection2.insert_many(data2)
print("Data inserted successfully!")

from sqlalchemy import create_engine
import os
from openpyxl import load_workbook

engine = create_engine("mysql+pymysql://root:Aassnnaa@localhost:3306/dwh_project")

# Example loading for multiple tables
csv_files = {
    'rides': 'rides_data.csv',
    'tickets': 'tickets.csv',
    'stations': 'station_data.csv',
    'ticket_office': 'ticket_office_data.csv',
    'vehicle': 'vehicle_data.csv'
}

for table_name, file_path in csv_files.items():
    print("hi")
    df = pd.read_csv(file_path)
    df.to_sql(table_name, con=engine, index=False, if_exists='replace')  
    print(f"Loaded {table_name} into MySQL.") 


files = [
    ('weather_data.csv', 'weather_data.xlsx'), 
    ('maintenance_data.csv', 'maintenance_data.xlsx'), 
    ('events_data.csv', 'events_data.xlsx'), 
    ('date_data.csv', 'date_data.xlsx'), 
    ('timeslot_data.csv', 'timeslot_data.xlsx')
    ]


filepath = '//wsl.localhost/Ubuntu/home/asnaf/airflow_project/project/data/excels'


def append_csv_to_excel(csv_filename, excel_filename):
    csv_path = csv_filename
    excel_path = os.path.join(filepath, excel_filename)

    # Read CSV
    new_data = pd.read_csv(csv_path)

    # Append or create Excel
    if os.path.exists(excel_path):
        book = load_workbook(excel_path)
        sheet = book.active
        start_row = sheet.max_row

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            new_data.to_excel(writer, index=False, header=False, startrow=start_row)
    else:
        new_data.to_excel(excel_path, index=False)

# Loop through all files
for csv_file, excel_file in files:
    append_csv_to_excel(csv_file, excel_file)

print("All CSVs processed and appended to Excel files successfully.")
# \\wsl.localhost\Ubuntu\home\asnaf\airflow_project\project\data\excels